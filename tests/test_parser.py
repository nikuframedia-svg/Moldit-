"""Tests for ISOP parser — Spec 01 §2, §7."""

from __future__ import annotations

import tempfile
from datetime import datetime

import pytest
from openpyxl import Workbook

from backend.parser.isop_reader import (
    _find_header_row,
    _safe_int,
    _safe_float,
    extract_stock_and_demand,
    read_isop,
)


# --- Helpers to build test ISOP workbooks ---


def _make_isop_wb(
    header_row: int = 5,
    has_twin: bool = False,
    rows: list[dict] | None = None,
    dates: list[datetime] | None = None,
) -> Workbook:
    """Create a minimal ISOP workbook for testing."""
    wb = Workbook()
    ws = wb.active

    if dates is None:
        dates = [datetime(2026, 3, 5), datetime(2026, 3, 6), datetime(2026, 3, 7)]

    # Header row
    headers = ["Cliente", "Nome", "Referência Artigo", "Designação",
               "Lote Económico", "Prz.Fabrico", "Máquina", "Ferramenta",
               "Peças/H", "Nº Pessoas", "STOCK-A", "WIP", "ATRASO"]
    if has_twin:
        headers.append("Peça Gémea")

    for col, h in enumerate(headers, 1):
        ws.cell(row=header_row, column=col, value=h)

    date_start = len(headers) + 1
    for i, d in enumerate(dates):
        ws.cell(row=header_row, column=date_start + i, value=d)

    # Data rows
    if rows is None:
        rows = [
            {
                "client_id": "210020", "client_name": "FAURECIA",
                "sku": "1064169X100", "designation": "Peça A",
                "eco_lot": 36400, "prz": "", "machine": "PRM031",
                "tool": "BFP079", "pH": 1681, "operators": 1,
                "stock_a": 5000, "wip": 0, "backlog": 0,
                "twin": "", "np": [2751, -15600, -10400],
            },
        ]

    for r_idx, row_data in enumerate(rows):
        r = header_row + 1 + r_idx
        vals = [
            row_data["client_id"], row_data["client_name"],
            row_data["sku"], row_data["designation"],
            row_data["eco_lot"], row_data.get("prz", ""),
            row_data["machine"], row_data["tool"],
            row_data["pH"], row_data["operators"],
            row_data.get("stock_a", 0), row_data["wip"], row_data["backlog"],
        ]
        if has_twin:
            vals.append(row_data.get("twin", ""))

        for col, v in enumerate(vals, 1):
            ws.cell(row=r, column=col, value=v)

        for i, np_val in enumerate(row_data["np"]):
            ws.cell(row=r, column=date_start + i, value=np_val if np_val != 0 else None)

    return wb


def _save_and_read(wb: Workbook) -> tuple:
    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as f:
        wb.save(f.name)
        wb.close()
        return read_isop(f.name)


# --- Test: extract_stock_and_demand ---


class TestStockAndDemand:
    def test_basic_stock_then_demand(self):
        """BFP079 example: stock=2751, then demands at day 7 and 11."""
        np = [2751, 2751, 2751, 2751, 2751, 0, 0, -15600, 0, 0, 0, -10400]
        stk, demand = extract_stock_and_demand(np)
        assert stk == 2751
        assert demand[7] == 15600
        assert demand[11] == 10400
        assert demand[0] == 0  # positive = stock, not demand
        assert sum(d for d in demand if d > 0) == 26000

    def test_no_stock_immediate_demand(self):
        np = [-5000, 0, -3000]
        stk, demand = extract_stock_and_demand(np)
        assert stk == 0
        assert demand == [5000, 0, 3000]

    def test_all_positive_no_demand(self):
        np = [100, 200, 300]
        stk, demand = extract_stock_and_demand(np)
        assert stk == 300  # last positive before first negative (no negative → last)
        assert demand == [0, 0, 0]

    def test_all_zeros(self):
        np = [0, 0, 0]
        stk, demand = extract_stock_and_demand(np)
        assert stk == 0
        assert demand == [0, 0, 0]

    def test_positive_after_negative_not_stock(self):
        """Positive values after first negative are NOT stock."""
        np = [100, -200, 50, -300]
        stk, demand = extract_stock_and_demand(np)
        assert stk == 100
        assert demand == [0, 200, 0, 300]


# --- Test: safe helpers ---


class TestSafeHelpers:
    def test_safe_int_normal(self):
        assert _safe_int(42) == 42
        assert _safe_int(42.7) == 42
        assert _safe_int("36400") == 36400

    def test_safe_int_edge(self):
        assert _safe_int(None) == 0
        assert _safe_int("") == 0
        assert _safe_int("abc") == 0

    def test_safe_float_normal(self):
        assert _safe_float(1681.0) == 1681.0
        assert _safe_float("1681") == 1681.0

    def test_safe_float_edge(self):
        assert _safe_float(None) == 0.0
        assert _safe_float("") == 0.0


# --- Test: header detection ---


class TestHeaderDetection:
    def test_header_at_row_5(self):
        wb = _make_isop_wb(header_row=5)
        ws = wb.active
        assert _find_header_row(ws) == 5

    def test_header_at_row_1(self):
        wb = _make_isop_wb(header_row=1)
        ws = wb.active
        assert _find_header_row(ws) == 1

    def test_header_not_found(self):
        wb = Workbook()
        ws = wb.active
        ws.cell(row=1, column=1, value="NotCliente")
        with pytest.raises(ValueError, match="Header row not found"):
            _find_header_row(ws)


# --- Test: read_isop integration ---


class TestReadIsop:
    def test_basic_read(self):
        wb = _make_isop_wb()
        rows, workdays, has_twin = _save_and_read(wb)
        assert len(rows) == 1
        assert len(workdays) == 3
        assert has_twin is False
        assert rows[0].sku == "1064169X100"
        assert rows[0].machine_id == "PRM031"
        assert rows[0].eco_lot == 36400

    def test_prm020_filtered(self):
        wb = _make_isop_wb(rows=[
            {
                "client_id": "X", "client_name": "TEST",
                "sku": "SKU1", "designation": "D",
                "eco_lot": 0, "machine": "PRM020",
                "tool": "T1", "pH": 100, "operators": 1,
                "wip": 0, "backlog": 0, "np": [0, -100, 0],
            },
            {
                "client_id": "Y", "client_name": "TEST2",
                "sku": "SKU2", "designation": "D2",
                "eco_lot": 0, "machine": "PRM031",
                "tool": "T2", "pH": 200, "operators": 1,
                "wip": 0, "backlog": 0, "np": [0, -200, 0],
            },
        ])
        rows, _, _ = _save_and_read(wb)
        assert len(rows) == 1
        assert rows[0].sku == "SKU2"

    def test_twin_column_detected(self):
        wb = _make_isop_wb(has_twin=True, rows=[
            {
                "client_id": "A", "client_name": "C1",
                "sku": "SKU_A", "designation": "D",
                "eco_lot": 0, "machine": "PRM031",
                "tool": "T1", "pH": 100, "operators": 1,
                "wip": 0, "backlog": 0, "twin": "SKU_B",
                "np": [-500, 0, 0],
            },
        ])
        rows, _, has_twin = _save_and_read(wb)
        assert has_twin is True
        assert rows[0].twin_ref == "SKU_B"

    def test_np_values_parsed(self):
        wb = _make_isop_wb(rows=[
            {
                "client_id": "A", "client_name": "C1",
                "sku": "SKU1", "designation": "D",
                "eco_lot": 1000, "machine": "PRM019",
                "tool": "T1", "pH": 500, "operators": 2,
                "wip": 10, "backlog": 50,
                "np": [2751, -15600, -10400],
            },
        ])
        rows, _, _ = _save_and_read(wb)
        assert rows[0].np_values == [2751, -15600, -10400]
        assert rows[0].wip == 10
        assert rows[0].backlog == 50
