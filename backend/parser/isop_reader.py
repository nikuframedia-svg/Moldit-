"""ISOP Excel parser — Spec 01 §2.

Reads ISOP Excel files with dynamic header detection and column mapping.
Supports multiple ISOP formats (completo, 17/03, 27/02).
"""

from __future__ import annotations

import logging
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

from backend.types import RawRow

logger = logging.getLogger(__name__)

# --- Column mapping by header name (NOT position) ---

COLUMN_MAP: dict[str, str] = {
    "Cliente": "client_id",
    "Nome": "client_name",
    "Produto Acabado": "produto_acabado",
    "Referência Artigo": "sku",
    "Designação": "designation",
    "Lote Económico": "eco_lot",
    "Máquina": "machine_id",
    "Máquina alternativa": "alt_machine",
    "Ferramenta": "tool_id",
    "Tp.Setup": "setup_hours",
    "Peças/H": "pieces_per_hour",
    "Nº Pessoas": "operators",
    "Pessoas": "operators",
    "Qtd Exp": "qty_exp",
    "WIP": "wip",
    "ATRASO": "backlog",
    "Peça Gémea": "twin_ref",
}

IGNORE: set[str] = {"Prz.Fabrico", "STOCK-A"}


# --- Header detection ---


def _find_header_row(ws) -> int:
    """Find the header row by scanning for 'Cliente' in column A."""
    for row in range(1, 21):
        val = ws.cell(row=row, column=1).value
        if val is not None and str(val).strip() == "Cliente":
            return row
    raise ValueError("Header row not found: no cell with 'Cliente' in column A (rows 1-20)")


# --- Dynamic column mapping ---


def _build_column_map(
    ws, header_row: int
) -> tuple[dict[str, int], int | None, bool]:
    """Build column index map from header names.

    Returns:
        (col_map, first_date_col, has_twin)
    """
    col_map: dict[str, int] = {}
    first_date_col: int | None = None
    has_twin = False

    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val is None:
            continue

        # Date columns: datetime objects or date-like values
        if isinstance(val, datetime):
            if first_date_col is None:
                first_date_col = col
            continue

        header = str(val).strip()
        if header in IGNORE:
            continue

        if header in COLUMN_MAP:
            field_name = COLUMN_MAP[header]
            col_map[field_name] = col
            if header == "Peça Gémea":
                has_twin = True

    return col_map, first_date_col, has_twin


# --- Date extraction ---


def _extract_dates(ws, header_row: int, first_date_col: int) -> list[str]:
    """Extract workday dates from header row (ISO format strings)."""
    dates: list[str] = []
    for col in range(first_date_col, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if isinstance(val, datetime):
            dates.append(val.strftime("%Y-%m-%d"))
        else:
            break
    return dates


# --- Safe value helpers ---


def _safe_int(val) -> int:
    if val is None or val == "":
        return 0
    try:
        return int(float(val))
    except (ValueError, TypeError):
        return 0


def _safe_float(val) -> float:
    if val is None or val == "":
        return 0.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0.0


def _get(ws, row: int, col_map: dict[str, int], field: str, default=None):
    """Get cell value by field name from column map."""
    col = col_map.get(field)
    if col is None:
        return default
    val = ws.cell(row=row, column=col).value
    return val if val is not None else default


# --- Stock and demand extraction ---


def extract_stock_and_demand(np_values: list[int]) -> tuple[int, list[int]]:
    """Extract stock and demand from NP values.

    Stock = last positive value before first negative.
    Demand = abs(negative values), 0 elsewhere.
    """
    stk = 0
    demand: list[int] = []
    found_negative = False

    for val in np_values:
        if val > 0 and not found_negative:
            stk = val
            demand.append(0)
        elif val < 0:
            found_negative = True
            demand.append(abs(val))
        else:
            demand.append(0)

    return stk, demand


# --- Main reader ---


def read_isop(path: str | Path) -> tuple[list[RawRow], list[str], bool]:
    """Read ISOP Excel file.

    Args:
        path: Path to .xlsx file.

    Returns:
        (rows, workdays, has_twin_column)
        - rows: list of RawRow (one per ISOP line, PRM020 filtered)
        - workdays: list of date strings ("2026-03-05")
        - has_twin_column: whether "Peça Gémea" column exists
    """
    wb = load_workbook(str(path), data_only=True)
    ws = wb.active

    header_row = _find_header_row(ws)
    col_map, first_date_col, has_twin = _build_column_map(ws, header_row)

    if first_date_col is None:
        wb.close()
        raise ValueError("No date columns found in ISOP header")

    workdays = _extract_dates(ws, header_row, first_date_col)
    n_dates = len(workdays)

    if n_dates == 0:
        wb.close()
        raise ValueError("No workdays extracted from ISOP header")

    rows: list[RawRow] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        sku = _get(ws, row_idx, col_map, "sku")
        if not sku or str(sku).strip() == "":
            break

        machine = str(_get(ws, row_idx, col_map, "machine_id", "")).strip()

        # PRM020 — FORA DE USO. IGNORAR SEMPRE.
        if machine == "PRM020":
            continue

        # Extract NP values from date columns
        np_values: list[int] = []
        for col in range(first_date_col, first_date_col + n_dates):
            v = ws.cell(row=row_idx, column=col).value
            np_values.append(_safe_int(v))

        # pH warning
        ph = _safe_float(_get(ws, row_idx, col_map, "pieces_per_hour", 0))
        if ph <= 0:
            logger.warning("pH=0 for SKU %s on %s, defaulting to 1.0", str(sku).strip(), machine)

        rows.append(
            RawRow(
                client_id=str(_get(ws, row_idx, col_map, "client_id", "")).strip(),
                client_name=str(_get(ws, row_idx, col_map, "client_name", "")).strip(),
                sku=str(sku).strip(),
                designation=str(_get(ws, row_idx, col_map, "designation", "")).strip(),
                eco_lot=_safe_int(_get(ws, row_idx, col_map, "eco_lot", 0)),
                machine_id=machine,
                tool_id=str(_get(ws, row_idx, col_map, "tool_id", "")).strip(),
                pieces_per_hour=ph,
                operators=_safe_int(_get(ws, row_idx, col_map, "operators", 1)),
                wip=_safe_int(_get(ws, row_idx, col_map, "wip", 0)),
                backlog=_safe_int(_get(ws, row_idx, col_map, "backlog", 0)),
                twin_ref=(
                    str(_get(ws, row_idx, col_map, "twin_ref", "")).strip()
                    if has_twin
                    else ""
                ),
                np_values=np_values,
            )
        )

    wb.close()

    logger.info(
        "Parsed ISOP: %d rows, %d workdays, twin_col=%s",
        len(rows),
        n_dates,
        has_twin,
    )

    return rows, workdays, has_twin
