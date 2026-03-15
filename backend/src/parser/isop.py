"""ISOP Excel parser — Contract C-01.

Reads the Nikufra ISOP XLSX and produces structured ISOPData.
Deterministic: same file always produces the same output.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path

import openpyxl

from src.engine.models import SKU, ISOPData, Order

# ── Column header patterns (case-insensitive) ──────────────────────

_COL_PATTERNS: dict[str, list[str]] = {
    "cliente": ["cliente"],
    "nome": ["nome"],
    "ref_artigo": [
        "referência artigo",
        "referencia artigo",
        "ref. artigo",
        "ref artigo",
    ],
    "designacao": ["designação", "designacao"],
    "lote_econ": ["lote econ", "lote económico", "lote economico"],
    "prz_fabrico": ["prz.fabrico", "prz fabrico", "prazo fabrico"],
    "maquina": ["máquina", "maquina"],
    "ferramenta": ["ferramenta"],
    "pecas_h": ["peças/h", "pecas/h", "pcs/h", "pçs/h", "cadência", "cadencia"],
    "n_pessoas": ["nº pessoas", "n pessoas", "num pessoas", "nº pess", "pessoas"],
    "stock_a": ["stock-a", "stock a"],
    "wip": ["wip"],
    "peca_gemea": ["peca gemea", "peça gémea", "peça gemea", "pç gemea", "twin"],
    "atraso": ["atraso"],
}

# Columns to IGNORE (per CLAUDE.md)
_IGNORE = {"prz_fabrico", "stock_a"}


# ── Helpers ─────────────────────────────────────────────────────────


def _norm(val: object) -> str:
    """Normalize a cell value to a stripped string."""
    if val is None:
        return ""
    return str(val).strip()


def _int_or(val: object, default: int = 0) -> int:
    """Parse a cell as int, returning *default* on failure."""
    if val is None:
        return default
    try:
        return int(float(str(val).replace(",", ".")))
    except (ValueError, TypeError):
        return default


def _find_header_row(ws: openpyxl.worksheet.worksheet.Worksheet) -> int | None:
    """Scan rows 1-16 for the header (must contain 'Referência Artigo' + 'Máquina')."""
    for row_idx in range(1, min(ws.max_row + 1, 17)):
        texts = [_norm(ws.cell(row=row_idx, column=c).value).lower() for c in range(1, ws.max_column + 1)]
        has_ref = any(
            "referência artigo" in t or "referencia artigo" in t or "ref. artigo" in t or "ref artigo" in t
            for t in texts
        )
        has_maq = any("máquina" in t or "maquina" in t for t in texts)
        if has_ref and has_maq:
            return row_idx
    return None


def _build_col_map(ws: openpyxl.worksheet.worksheet.Worksheet, header_row: int) -> dict[str, int]:
    """Map logical field names → 1-based column indices from the header row."""
    headers: list[str] = []
    for c in range(1, ws.max_column + 1):
        headers.append(_norm(ws.cell(row=header_row, column=c).value).lower())

    col_map: dict[str, int] = {}
    for field, patterns in _COL_PATTERNS.items():
        for ci, h in enumerate(headers):
            if any(p in h for p in patterns):
                col_map[field] = ci + 1  # 1-based
                break
    return col_map


def _parse_date(val: object) -> date | None:
    """Try to interpret a header cell as a date."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # Try Excel serial number
    try:
        serial = int(float(s))
        if 40000 < serial < 60000:
            from openpyxl.utils.datetime import from_excel

            return from_excel(serial).date()
    except (ValueError, TypeError):
        pass
    return None


def _detect_date_columns(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    col_map: dict[str, int],
) -> list[tuple[int, date]]:
    """Find date columns (after the last fixed column) and return (col_idx, date) pairs."""
    # Start searching after the last known fixed column
    last_fixed = max(col_map.values()) if col_map else 10
    start_col = last_fixed + 1

    date_cols: list[tuple[int, date]] = []
    for c in range(start_col, ws.max_column + 1):
        d = _parse_date(ws.cell(row=header_row, column=c).value)
        if d is not None:
            date_cols.append((c, d))

    # Fallback: try from column 10 onwards if nothing found
    if not date_cols:
        for c in range(10, ws.max_column + 1):
            d = _parse_date(ws.cell(row=header_row, column=c).value)
            if d is not None:
                date_cols.append((c, d))

    return date_cols


def _detect_twins_by_tool(skus: dict[str, "SKU"]) -> set[tuple[str, str]]:
    """Detect twin pairs from SKUs sharing same tool + machine (LH/RH pattern).

    When the ISOP has no explicit twin column, twins are detected by:
    - Same tool AND same machine
    - Exactly 2 SKUs per tool (not 1 or 3+)
    """
    from collections import defaultdict

    tool_machine_groups: dict[tuple[str, str], list[str]] = defaultdict(list)  # type: ignore[assignment]
    for sku in skus.values():
        if sku.tool:
            tool_machine_groups[(sku.tool, sku.machine)].append(sku.sku)

    pairs: set[tuple[str, str]] = set()
    for group_skus in tool_machine_groups.values():
        if len(group_skus) == 2:
            pair = tuple(sorted(group_skus))
            pairs.add(pair)  # type: ignore[arg-type]
    return pairs


def _extract_stock(np_values: list[float | None]) -> int:
    """Stock = last positive value before the first negative. 0 if all negative or empty."""
    last_positive: int = 0
    for v in np_values:
        if v is None:
            continue
        if v < 0:
            # Hit first negative — return what we accumulated
            return last_positive
        last_positive = int(v)
    # No negatives found — return last positive (or 0)
    return last_positive


# ── Main parser ─────────────────────────────────────────────────────


def parse_isop(path: Path) -> ISOPData:
    """Parse an ISOP Excel file and return structured ISOPData.

    Deterministic: same file always produces the same output.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # 1. Find header row
    header_row = _find_header_row(ws)
    if header_row is None:
        msg = "Header row not found (expected 'Referência Artigo' + 'Máquina' in rows 1-16)"
        raise ValueError(msg)

    # 2. Build column map
    col_map = _build_col_map(ws, header_row)
    if "ref_artigo" not in col_map or "maquina" not in col_map:
        msg = "Required columns 'Referência Artigo' and 'Máquina' not found in headers"
        raise ValueError(msg)

    # 3. Detect date columns
    date_cols = _detect_date_columns(ws, header_row, col_map)
    if not date_cols:
        msg = "No date columns found in the header row"
        raise ValueError(msg)

    # 4. Parse data rows
    all_orders: list[Order] = []
    skus: dict[str, SKU] = {}
    twin_set: set[tuple[str, str]] = set()
    machines_set: set[str] = set()
    tools_set: set[str] = set()

    def _cell(row: int, field: str) -> object:
        ci = col_map.get(field)
        if ci is None:
            return None
        return ws.cell(row=row, column=ci).value

    data_start = header_row + 1
    for row_idx in range(data_start, ws.max_row + 1):
        sku_raw = _norm(_cell(row_idx, "ref_artigo"))
        if not sku_raw:
            continue

        machine = _norm(_cell(row_idx, "maquina"))
        if not machine:
            continue

        tool = _norm(_cell(row_idx, "ferramenta"))
        client_code = _norm(_cell(row_idx, "cliente"))
        client_name = _norm(_cell(row_idx, "nome"))
        designation = _norm(_cell(row_idx, "designacao")) or sku_raw
        pieces_per_hour = _int_or(_cell(row_idx, "pecas_h"), 0)
        operators = _int_or(_cell(row_idx, "n_pessoas"), 1)
        economic_lot = _int_or(_cell(row_idx, "lote_econ"), 0)
        wip = _int_or(_cell(row_idx, "wip"), 0)  # noqa: F841 — kept for future use
        atraso = _int_or(_cell(row_idx, "atraso"), 0)

        twin_raw = _norm(_cell(row_idx, "peca_gemea"))
        twin_ref = twin_raw if twin_raw else None

        # Read NP values from date columns
        np_values: list[float | None] = []
        for col_idx, _dt in date_cols:
            raw = ws.cell(row=row_idx, column=col_idx).value
            if raw is None:
                np_values.append(None)
                continue
            s = str(raw).strip().replace(",", ".")
            if s == "":
                np_values.append(None)
                continue
            try:
                np_values.append(float(s))
            except (ValueError, TypeError):
                np_values.append(None)

        # Extract stock (last positive before first negative)
        stock = _extract_stock(np_values)

        # Create orders from negative NP cells
        row_orders: list[Order] = []
        for i, v in enumerate(np_values):
            if v is not None and v < 0:
                order = Order(
                    sku=sku_raw,
                    client_code=client_code,
                    client_name=client_name,
                    qty=abs(int(v)),
                    deadline=date_cols[i][1],
                    tool=tool,
                    machine=machine,
                    pieces_per_hour=pieces_per_hour,
                    operators=operators,
                    economic_lot=economic_lot,
                    twin_ref=twin_ref,
                )
                row_orders.append(order)

        all_orders.extend(row_orders)
        machines_set.add(machine)
        if tool:
            tools_set.add(tool)

        # Twin pair tracking
        if twin_ref:
            pair = tuple(sorted([sku_raw, twin_ref]))
            twin_set.add(pair)  # type: ignore[arg-type]

        # Aggregate into SKU
        if sku_raw in skus:
            existing = skus[sku_raw]
            existing.orders.extend(row_orders)
            if client_code and client_code not in existing.clients:
                existing.clients.append(client_code)
            # Keep worst (most negative) atraso
            if atraso < existing.atraso:
                skus[sku_raw] = existing.model_copy(update={"atraso": atraso})
        else:
            skus[sku_raw] = SKU(
                sku=sku_raw,
                designation=designation,
                machine=machine,
                tool=tool,
                pieces_per_hour=pieces_per_hour,
                operators=operators,
                economic_lot=economic_lot,
                twin_ref=twin_ref,
                stock=stock,
                atraso=atraso,
                orders=row_orders,
                clients=[client_code] if client_code else [],
            )

    # 5. Auto-detect twin pairs from shared tool when no explicit column
    if not twin_set:
        twin_set = _detect_twins_by_tool(skus)
        # Back-fill twin_ref on SKUs
        for a, b in twin_set:
            if a in skus and skus[a].twin_ref is None:
                skus[a] = skus[a].model_copy(update={"twin_ref": b})
            if b in skus and skus[b].twin_ref is None:
                skus[b] = skus[b].model_copy(update={"twin_ref": a})

    # 6. Build output
    all_dates = [d for _, d in date_cols]
    workdays = [d for d in all_dates if d.weekday() < 5]  # Mon=0 .. Fri=4

    twin_pairs = sorted(twin_set)

    return ISOPData(
        skus=skus,
        orders=all_orders,
        machines=sorted(machines_set),
        tools=sorted(tools_set),
        twin_pairs=twin_pairs,
        date_range=(min(all_dates), max(all_dates)),
        workdays=workdays,
    )
