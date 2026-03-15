"""ISOP Excel Parser — reads Incompol's real format."""
import openpyxl
from datetime import datetime
from models import Reference, Machine


def parse_isop(filepath: str) -> tuple[list[Reference], list[Machine], dict]:
    """Parse ISOP Excel into References and Machines.

    Returns:
        (references, machines, metadata)
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[wb.sheetnames[0]]

    # Find header row (row with 'Referência Artigo')
    header_row = None
    for row_idx in range(1, min(10, ws.max_row + 1)):
        for col_idx in range(1, min(20, ws.max_column + 1)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val and 'Refer' in str(val) and 'Artigo' in str(val):
                header_row = row_idx
                break
        if header_row:
            break

    if not header_row:
        raise ValueError("Could not find header row in ISOP file")

    # Map date columns (from column O onwards)
    date_columns = []
    for col_idx in range(15, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col_idx).value
        if val and isinstance(val, datetime):
            date_columns.append((col_idx, val.strftime("%Y-%m-%d")))

    # Parse references
    references = []
    machines_set = set()
    seen_refs = {}  # Track ref+client combos to aggregate demands

    for row_idx in range(header_row + 1, ws.max_row + 1):
        ref_val = ws.cell(row=row_idx, column=3).value
        if not ref_val or ref_val == 'Referência Artigo':
            continue

        ref_id = str(ref_val).strip()
        client_code = str(ws.cell(row=row_idx, column=1).value or "").strip()
        client_name = str(ws.cell(row=row_idx, column=2).value or "").strip()
        designation = str(ws.cell(row=row_idx, column=4).value or "").strip()
        economic_lot = int(ws.cell(row=row_idx, column=5).value or 0)
        lead_time = int(ws.cell(row=row_idx, column=6).value or 0)
        machine = str(ws.cell(row=row_idx, column=7).value or "").strip()
        tool = str(ws.cell(row=row_idx, column=8).value or "").strip()
        pcs_hour = int(ws.cell(row=row_idx, column=9).value or 0)
        num_people = int(ws.cell(row=row_idx, column=10).value or 1)
        stock = int(ws.cell(row=row_idx, column=11).value or 0)
        wip_val = ws.cell(row=row_idx, column=12).value
        wip = int(wip_val) if wip_val and str(wip_val).strip() else 0
        twin_val = ws.cell(row=row_idx, column=13).value
        twin_ref = str(twin_val).strip() if twin_val and str(twin_val).strip() else None

        # Daily coverage data
        daily_coverage = {}
        for col_idx, date_str in date_columns:
            val = ws.cell(row=row_idx, column=col_idx).value
            if val is not None and isinstance(val, (int, float)):
                daily_coverage[date_str] = int(val)

        if machine:
            machines_set.add(machine)

        # Unique key is ref+client
        key = f"{ref_id}_{client_code}"

        ref = Reference(
            ref_id=ref_id,
            client_code=client_code,
            client_name=client_name,
            designation=designation,
            economic_lot=economic_lot,
            lead_time_days=lead_time,
            machine=machine,
            tool=tool,
            pieces_per_hour=pcs_hour,
            num_people=num_people,
            stock=stock,
            wip=wip,
            twin_ref=twin_ref,
            daily_coverage=daily_coverage,
        )
        references.append(ref)

    # Create machine objects
    machines = []
    for m in sorted(machines_set):
        machines.append(Machine(machine_id=m))

    # Metadata
    meta_cell = ws.cell(row=1, column=1).value or ""
    metadata = {
        "total_refs": len(references),
        "total_machines": len(machines),
        "date_range": f"{date_columns[0][1]} to {date_columns[-1][1]}" if date_columns else "",
        "header_info": str(meta_cell),
        "parse_date": datetime.now().isoformat(),
    }

    return references, machines, metadata


if __name__ == "__main__":
    import sys
    filepath = sys.argv[1] if len(sys.argv) > 1 else "/mnt/user-data/uploads/ISOP__Nikufra_27_2-2.xlsx"
    refs, machines, meta = parse_isop(filepath)
    print(f"Parsed {len(refs)} references, {len(machines)} machines")
    print(f"Machines: {[m.machine_id for m in machines]}")
    print(f"Date range: {meta['date_range']}")

    # Show urgency
    urgent = [r for r in refs if r.first_shortage_date]
    print(f"\nRefs with shortage: {len(urgent)}")
    for r in sorted(urgent, key=lambda x: x.first_shortage_date or "9999")[:10]:
        print(f"  {r.ref_id} ({r.client_name}) | Maq={r.machine} | Stock={r.stock} | Ruptura={r.first_shortage_date}")
