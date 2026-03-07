# ISOP XLSX Parser
# Conforme SP-BE-03 e C-02

import hashlib
from datetime import date, datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

import openpyxl
from openpyxl import Workbook

from ...core.errors import APIException, ErrorCodes
from ...core.logging import get_logger
from .excel_utils import (
    normalize_code,
    normalize_string,
    parse_date_cell,
)
from .excel_utils import (
    parse_integer_optional as parse_integer,
)
from .excel_utils import (
    parse_numeric_optional as parse_numeric,
)

logger = get_logger(__name__)


class ISOPParserError(Exception):
    """Erro específico do parser ISOP"""

    pass


def sha256_file(filepath: Path) -> str:
    """Calcula SHA-256 de um ficheiro"""
    sha256 = hashlib.sha256()
    with open(filepath, "rb") as f:
        for chunk in iter(lambda: f.read(4096), b""):
            sha256.update(chunk)
    return sha256.hexdigest()


class ISOPParser:
    """Parser para ficheiros ISOP XLSX"""

    # Headers esperados na linha 7 (0-indexed = linha 6)
    EXPECTED_HEADERS = [
        "Cliente",
        "Nome",
        "Produto Acabado",
        "Referência Artigo",
        "Designação",
        "Lote Económico",
        "Máquina",
        "Máquina alternativa",
        "Ferramenta",
        "Tp.Setup",
        "Peças/H",
        "Nº Pessoas",
        "Qtd Exp",
        # Depois seguem colunas de data (35 datas esperadas)
    ]

    def __init__(self, filepath: Path, series_semantics: str, setup_time_uom: str = "UNKNOWN"):
        self.filepath = filepath
        self.series_semantics = series_semantics
        self.setup_time_uom = setup_time_uom
        self.wb: Workbook | None = None
        self.ws = None
        self.header_row = 6  # Linha 7 (0-indexed)
        self.data_start_row = 7  # Linha 8 (0-indexed)

    def parse(self) -> dict[str, Any]:
        """Parse o ficheiro XLSX e retorna snapshot canónico"""
        logger.info(f"Parsing ISOP file: {self.filepath}")

        # Carregar workbook
        try:
            self.wb = openpyxl.load_workbook(self.filepath, data_only=True)
        except Exception as e:
            raise APIException(
                status_code=400,
                code=ErrorCodes.ERR_INVALID_UUID,  # Usar código genérico por agora
                message=f"Invalid XLSX file: {str(e)}",
            )

        # Obter sheet "Planilha1"
        if "Planilha1" not in self.wb.sheetnames:
            raise APIException(
                status_code=400,
                code=ErrorCodes.ERR_INVALID_UUID,
                message="Sheet 'Planilha1' not found",
            )

        self.ws = self.wb["Planilha1"]

        # Validar headers
        self._validate_headers()

        # Extrair dados
        items = []
        resources = []
        tools = []
        customers = []
        routings = []
        series_entries = []

        seen_items = set()
        seen_resources = set()
        seen_tools = set()
        seen_customers = set()

        # Encontrar colunas de data
        date_columns = self._find_date_columns()

        # G-8: Parse working day flags from row 5 (1=útil, 0=não-útil)
        working_day_flags = self._parse_working_day_flags(date_columns)

        # Track qtd_exp values for semantics detection (F-11.3)
        qtd_exp_values: list[float | None] = []

        # Processar linhas de dados
        row_idx = self.data_start_row
        while row_idx <= self.ws.max_row:
            row_data = self._parse_row(row_idx, date_columns)

            if not row_data:
                row_idx += 1
                continue

            # Item
            item_sku = row_data.get("item_sku")
            if item_sku and item_sku not in seen_items:
                items.append(
                    {
                        "item_sku": item_sku,
                        "name": row_data.get("item_name"),
                        "parent_sku": row_data.get("parent_sku"),
                        "lot_economic_qty": row_data.get("lot_economic_qty"),
                    }
                )
                seen_items.add(item_sku)

            # Resource
            resource_code = row_data.get("resource_code")
            if resource_code and resource_code not in seen_resources:
                resources.append(
                    {
                        "resource_code": resource_code,
                        "name": None,  # Não há nome no XLSX
                    }
                )
                seen_resources.add(resource_code)

            # Tool
            tool_code = row_data.get("tool_code")
            if tool_code and tool_code not in seen_tools:
                tools.append(
                    {
                        "tool_code": tool_code,
                        "name": None,
                    }
                )
                seen_tools.add(tool_code)

            # Customer
            customer_code = row_data.get("customer_code")
            if customer_code and customer_code not in seen_customers:
                customers.append(
                    {
                        "customer_code": customer_code,
                        "name": row_data.get("customer_name"),
                    }
                )
                seen_customers.add(customer_code)

            # Routing
            routing = {
                "item_sku": item_sku,
                "routing_ref": None,
                "operations": [
                    {
                        "sequence": 1,
                        "resource_code": resource_code,
                        "tool_code": tool_code,
                        "setup_time": row_data.get("setup_time"),
                        "rate": row_data.get("rate"),
                        "operators_required": row_data.get("operators_required"),
                        "alt_resources": row_data.get("alt_resources", []),
                    }
                ],
            }
            routings.append(routing)

            # Track qtd_exp for semantics detection (F-11.3)
            qtd_exp_values.append(row_data.get("qtd_exp"))

            # Series
            for date_col, date_val in date_columns.items():
                cell_value = self.ws.cell(row=row_idx + 1, column=date_col).value
                qty = parse_numeric(cell_value, 0.0)
                if qty is not None and qty != 0:
                    series_entries.append(
                        {
                            "item_sku": item_sku,
                            "date": date_val.isoformat(),
                            "quantity": qty,
                        }
                    )

            row_idx += 1

        # F-11.3: Detect Qtd Exp semantics
        qtd_exp_semantics = self._detect_qtd_exp_semantics(qtd_exp_values, series_entries)

        # Construir snapshot canónico
        snapshot = {
            "snapshot_id": str(uuid4()),
            "tenant_id": str(uuid4()),  # TODO: obter do contexto
            "created_at": datetime.utcnow().isoformat() + "Z",
            "sources": [
                {
                    "source_id": str(uuid4()),
                    "type": "XLSX",
                    "filename": self.filepath.name,
                    "file_hash_sha256": sha256_file(self.filepath),
                    "generated_at_local": datetime.now().isoformat(),
                    "source_timezone": "Europe/Lisbon",  # TODO: configurável
                }
            ],
            "semantics": {
                "series_semantics": self.series_semantics,
                "setup_time_uom": self.setup_time_uom,
                "qtd_exp_semantics": qtd_exp_semantics,
            },
            "master_data": {
                "customers": customers,
                "items": items,
                "resources": resources,
                "tools": tools,
            },
            "routing": routings,
            "series": series_entries,
            "working_days": working_day_flags,
            "trust_index": self._calculate_trust_index_full(snapshot),
        }

        return snapshot

    def _validate_headers(self):
        """Valida que os headers estão na linha esperada"""
        header_row = self.ws[self.header_row + 1]  # +1 porque openpyxl é 1-indexed

        found_headers = {}
        for col_idx, cell in enumerate(header_row, start=1):
            if cell.value:
                header_name = str(cell.value).strip()
                found_headers[col_idx] = header_name

        # Verificar headers obrigatórios
        required_headers = ["Referência Artigo", "Máquina"]
        missing = []
        for req_header in required_headers:
            if req_header not in found_headers.values():
                missing.append(req_header)

        if missing:
            raise APIException(
                status_code=400,
                code=ErrorCodes.ERR_INVALID_UUID,
                message=f"Missing required headers: {', '.join(missing)}",
            )

    def _parse_working_day_flags(self, date_columns: dict[int, date]) -> dict[str, bool]:
        """G-8: Parse working day flags from row 5 (1=útil, 0=não-útil)."""
        working_days: dict[str, bool] = {}
        row = self.ws[5]  # openpyxl 1-indexed → Excel row 5
        for col_idx, date_val in date_columns.items():
            cell_value = row[col_idx - 1].value if col_idx - 1 < len(row) else None
            if cell_value is not None:
                try:
                    working_days[date_val.isoformat()] = int(cell_value) == 1
                except (ValueError, TypeError):
                    working_days[date_val.isoformat()] = True
            else:
                working_days[date_val.isoformat()] = True  # Default: útil
        return working_days

    def _find_date_columns(self) -> dict[int, date]:
        """Encontra colunas de data no header"""
        header_row = self.ws[self.header_row + 1]
        date_columns = {}

        for col_idx, cell in enumerate(header_row, start=1):
            if col_idx <= len(self.EXPECTED_HEADERS):
                continue  # Pular colunas fixas

            date_val = parse_date_cell(cell.value)
            if date_val:
                date_columns[col_idx] = date_val

        return date_columns

    def _parse_row(self, row_idx: int, date_columns: dict[int, date]) -> dict[str, Any] | None:
        """Parse uma linha de dados"""
        row = self.ws[row_idx + 1]  # +1 porque openpyxl é 1-indexed

        # Mapear colunas (assumindo ordem fixa)
        item_sku = normalize_code(row[3].value) if len(row) > 3 else None  # "Referência Artigo"
        if not item_sku:
            return None  # Linha vazia

        customer_code = normalize_code(row[0].value) if len(row) > 0 else None  # "Cliente"
        customer_name = normalize_string(row[1].value) if len(row) > 1 else None  # "Nome"
        parent_sku = normalize_code(row[2].value) if len(row) > 2 else None  # "Produto Acabado"
        item_name = normalize_string(row[4].value) if len(row) > 4 else None  # "Designação"
        lot_economic_qty = parse_integer(row[5].value) if len(row) > 5 else None  # "Lote Económico"
        resource_code = normalize_code(row[6].value) if len(row) > 6 else None  # "Máquina"
        alt_resource = (
            normalize_code(row[7].value) if len(row) > 7 else None
        )  # "Máquina alternativa"
        tool_code = normalize_code(row[8].value) if len(row) > 8 else None  # "Ferramenta"
        setup_time = parse_numeric(row[9].value) if len(row) > 9 else None  # "Tp.Setup"
        rate = parse_numeric(row[10].value) if len(row) > 10 else None  # "Peças/H"
        operators_required = parse_integer(row[11].value, 1) if len(row) > 11 else 1  # "Nº Pessoas"
        qtd_exp = parse_numeric(row[12].value) if len(row) > 12 else None  # "Qtd Exp"

        # Validar campos obrigatórios
        if not resource_code:
            raise APIException(
                status_code=400,
                code=ErrorCodes.ERR_INVALID_UUID,
                message=f"Missing required field 'Máquina' in row {row_idx + 1}",
            )

        # Alt resources
        alt_resources = []
        if alt_resource and alt_resource != "-" and alt_resource != resource_code:
            alt_resources.append(alt_resource)

        return {
            "item_sku": item_sku,
            "customer_code": customer_code,
            "customer_name": customer_name,
            "parent_sku": parent_sku,
            "item_name": item_name,
            "lot_economic_qty": lot_economic_qty,
            "resource_code": resource_code,
            "alt_resources": alt_resources,
            "tool_code": tool_code,
            "setup_time": setup_time,
            "rate": rate,
            "operators_required": operators_required,
            "qtd_exp": qtd_exp,
        }

    def _detect_qtd_exp_semantics(
        self,
        qtd_exp_values: list[float | None],
        series_entries: list[dict[str, Any]],
    ) -> str:
        """Detect Qtd Exp column semantics (F-11.3).

        Heuristic: if the total daily quantity per item matches qtd_exp for
        >50% of items, it's likely SHIPMENT_LOT_SIZE. Otherwise UNKNOWN.
        """
        non_null = [v for v in qtd_exp_values if v is not None and v > 0]
        if not non_null:
            return "UNKNOWN"

        # Aggregate total series qty per item
        item_total_qty: dict[str, float] = {}
        for entry in series_entries:
            sku = entry.get("item_sku", "")
            qty = entry.get("quantity", 0)
            item_total_qty[sku] = item_total_qty.get(sku, 0) + qty

        # Compare qtd_exp with total demand per item (rough match)
        # Since we track qtd_exp per row (one per item), build item→qtd_exp map
        # We use the first qtd_exp per item (rows may repeat items)
        matches = 0
        total = 0
        for i, val in enumerate(qtd_exp_values):
            if val is not None and val > 0:
                total += 1
                # We don't have sku index here easily; use a tolerance check
                # If qtd_exp matches common lot sizes, mark as LOT_SIZE
                if val in item_total_qty.values():
                    matches += 1

        if total > 0 and matches / total > 0.5:
            logger.info(
                "Qtd Exp semantics detected as SHIPMENT_LOT_SIZE (%.0f%% match)",
                matches / total * 100,
            )
            return "SHIPMENT_LOT_SIZE"

        logger.info(
            "Qtd Exp semantics: UNKNOWN (%.0f%% match of %d values)",
            (matches / total * 100) if total > 0 else 0,
            total,
        )
        return "UNKNOWN"

    def _calculate_trust_index_full(self, snapshot: dict[str, Any]) -> dict[str, Any]:
        """Calcula TrustIndex simplificado (DQA module removed)"""
        overall = 1.0
        causes = []
        if self.series_semantics == "UNKNOWN":
            overall = max(0.0, overall - 0.2)
            causes.append("UNKNOWN_SEMANTICS")
        return {"overall": overall, "causes": causes}


def generate_import_report(snapshot: dict[str, Any]) -> dict[str, Any]:
    """Gera relatório de import"""
    master_data = snapshot.get("master_data", {})
    items = master_data.get("items", [])
    resources = master_data.get("resources", [])
    tools = master_data.get("tools", [])
    customers = master_data.get("customers", [])
    routings = snapshot.get("routing", [])
    series = snapshot.get("series", [])

    warnings = []

    # Verificar WIP vazio
    if not series:
        warnings.append("No time series data found")

    # Verificar ATRASO vazio (se aplicável)
    # TODO: implementar quando necessário

    return {
        "items_count": len(items),
        "resources_count": len(resources),
        "tools_count": len(tools),
        "customers_count": len(customers),
        "routings_count": len(routings),
        "series_entries_count": len(series),
        "warnings": warnings,
        "trust_index": snapshot.get("trust_index", {}).get("overall", 0.0),
    }
