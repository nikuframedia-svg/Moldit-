"""ISOP XLSX -> NikufraData parser (Python port of frontend isop/ 923 LOC).

Mirrors the TypeScript parser logic:
  - Dynamic header detection (scan rows 0-15)
  - Flexible column mapping by pattern matching
  - Red cell highlighting detection (openpyxl styles)
  - Raw NP values in operations[].d
  - Trust score computation
  - Workday flag extraction

Output: NikufraData-compatible dict ready for scheduling pipeline.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import date
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from .column_mapper import (
    build_column_map,
    find_date_columns,
    find_header_row,
    find_workday_flags_row,
)
from .constants import DAY_NAMES_PT
from .helpers import _format_date
from .row_extractor import ParsedRow, extract_rows

logger = logging.getLogger(__name__)

# ── Machine -> Area mapping ──────────────────────────────────────

MACHINE_AREA: dict[str, str] = {
    "PRM019": "PG1",
    "PRM020": "PG1",
    "PRM043": "PG1",
    "PRM031": "PG2",
    "PRM039": "PG2",
    "PRM042": "PG2",
}


# ── Trust score result ──────────────────────────────────────────


@dataclass
class TrustScoreResult:
    score: float = 0.0
    dimensions: dict[str, float] = field(
        default_factory=lambda: {
            "completeness": 0.0,
            "quality": 0.0,
            "demandCoverage": 0.0,
            "consistency": 0.0,
        }
    )


# ── Parse result ────────────────────────────────────────────────


@dataclass
class ISOPParseResult:
    success: bool = True
    data: dict[str, Any] | None = None  # NikufraData dict
    meta: dict[str, Any] | None = None
    source_columns: dict[str, bool] | None = None
    errors: list[str] = field(default_factory=list)


# ═══════════════════════════════════════════════════════════════
#  Trust score
# ═══════════════════════════════════════════════════════════════


def _compute_trust_score(
    rows: list[ParsedRow],
    tools: list[dict[str, Any]],
    operations: list[dict[str, Any]],
) -> TrustScoreResult:
    if not rows:
        return TrustScoreResult()

    n = len(rows)

    # 1. Completeness (40%)
    complete = sum(
        1
        for r in rows
        if r.item_sku and r.resource_code and r.tool_code and r.rate > 0 and r.setup_time >= 0
    )
    completeness = complete / n

    # 2. Quality (30%)
    valid = sum(1 for r in rows if r.rate > 0 and r.setup_time >= 0 and r.operators_required >= 1)
    quality = valid / n

    # 3. Demand coverage (20%)
    with_demand = sum(
        1 for op in operations if any(v is not None and v != 0 for v in op.get("d", []))
    )
    demand_coverage = with_demand / len(operations) if operations else 0.0

    # 4. Consistency (10%)
    machine_set = {r.resource_code for r in rows}
    valid_tools = sum(1 for t in tools if t.get("m") in machine_set)
    consistency = valid_tools / len(tools) if tools else 1.0

    score = round(
        completeness * 0.4 + quality * 0.3 + demand_coverage * 0.2 + consistency * 0.1,
        2,
    )

    return TrustScoreResult(
        score=score,
        dimensions={
            "completeness": round(completeness, 2),
            "quality": round(quality, 2),
            "demandCoverage": round(demand_coverage, 2),
            "consistency": round(consistency, 2),
        },
    )


# ═══════════════════════════════════════════════════════════════
#  Build NikufraData
# ═══════════════════════════════════════════════════════════════


def _day_label(d: date) -> str:
    return DAY_NAMES_PT[d.weekday()]


def _build_nikufra_data(
    parsed_rows: list[ParsedRow],
    dates: list[date],
    workday_flags: list[bool],
    warnings: list[str],
    source_columns: dict[str, bool],
) -> ISOPParseResult:
    n_days = len(dates)

    # ── Machines ──
    machine_set: dict[str, str] = {}  # id -> area
    for r in parsed_rows:
        if r.resource_code not in machine_set:
            machine_set[r.resource_code] = MACHINE_AREA.get(r.resource_code, "PG1")
        if r.alt_resource and r.alt_resource not in machine_set:
            machine_set[r.alt_resource] = MACHINE_AREA.get(r.alt_resource, "PG1")

    machines_down = {r.resource_code for r in parsed_rows if r.machine_down}
    if machines_down:
        warnings.append(
            "Máquinas inoperacionais detectadas (texto/cor): " + ", ".join(sorted(machines_down))
        )

    unknown_machines = [mid for mid in machine_set if mid not in MACHINE_AREA]
    if unknown_machines:
        warnings.append(
            "Máquina(s) desconhecida(s) atribuída(s) a PG1 por defeito: "
            + ", ".join(unknown_machines)
        )

    machines = []
    for mid in sorted(machine_set):
        m: dict[str, Any] = {
            "id": mid,
            "area": machine_set[mid],
            "man": [0] * n_days,
        }
        if mid in machines_down:
            m["status"] = "down"
        machines.append(m)

    # ── Tools ──
    tool_map: dict[str, dict[str, Any]] = {}
    for row in parsed_rows:
        if not row.tool_code:
            continue
        existing = tool_map.get(row.tool_code)
        if existing:
            if row.item_sku not in existing["skus"]:
                existing["skus"].append(row.item_sku)
                existing["nm"].append(row.item_name)
            if row.resource_code != existing["m"]:
                warnings.append(
                    f'Ferramenta "{row.tool_code}" aparece com máquinas diferentes: '
                    f"{existing['m']} (mantida) vs {row.resource_code} (SKU {row.item_sku})"
                )
            existing["wip"] = max(existing.get("wip", 0), row.wip)
        else:
            tool_map[row.tool_code] = {
                "id": row.tool_code,
                "m": row.resource_code,
                "alt": row.alt_resource or "-",
                "s": row.setup_time,
                "pH": row.rate,
                "op": row.operators_required,
                "skus": [row.item_sku],
                "nm": [row.item_name],
                "lt": row.lot_economic_qty,
                "stk": 0,
                "wip": row.wip,
            }

    tools_down = {r.tool_code for r in parsed_rows if r.tool_down and r.tool_code}
    if tools_down:
        warnings.append(
            "Ferramentas inoperacionais detectadas (texto/cor): " + ", ".join(sorted(tools_down))
        )

    tools = list(tool_map.values())
    for t in tools:
        if t["id"] in tools_down:
            t["status"] = "down"

    # ── Customers ──
    customer_map: dict[str, str] = {}
    for r in parsed_rows:
        if r.customer_code and r.customer_code not in customer_map:
            customer_map[r.customer_code] = r.customer_name
    customers = [{"code": code, "name": name} for code, name in sorted(customer_map.items())]

    # ── Operations ──
    ops_without_tool = [r for r in parsed_rows if not r.tool_code]
    if ops_without_tool:
        skus = ", ".join(r.item_sku for r in ops_without_tool[:5])
        suffix = "..." if len(ops_without_tool) > 5 else ""
        warnings.append(
            f"{len(ops_without_tool)} operação(ões) sem código de ferramenta — "
            f"não serão agendadas: {skus}{suffix}"
        )

    operations: list[dict[str, Any]] = []
    for idx, row in enumerate(parsed_rows):
        op: dict[str, Any] = {
            "id": f"OP{idx + 1:02d}",
            "m": row.resource_code,
            "t": row.tool_code,
            "sku": row.item_sku,
            "nm": row.item_name,
            "pH": row.rate,
            "atr": row.atraso,
            "d": row.daily_quantities,
            "s": row.setup_time,
            "op": row.operators_required,
        }
        if row.customer_code:
            op["cl"] = row.customer_code
        if row.customer_name:
            op["clNm"] = row.customer_name
        if row.parent_sku:
            op["pa"] = row.parent_sku
        if row.wip:
            op["wip"] = row.wip
        if row.qtd_exp:
            op["qe"] = row.qtd_exp
        if row.lead_time_days:
            op["ltDays"] = row.lead_time_days
        if row.twin:
            op["twin"] = row.twin
        operations.append(op)

    # ── MO load ──
    mo = {"PG1": [0.0] * n_days, "PG2": [0.0] * n_days}

    # ── Date labels ──
    date_labels = [_format_date(d) for d in dates]
    day_labels = [_day_label(d) for d in dates]

    nikufra_data: dict[str, Any] = {
        "dates": date_labels,
        "days_label": day_labels,
        "mo": mo,
        "machines": machines,
        "tools": tools,
        "operations": operations,
        "history": [],
        "customers": customers,
        "workday_flags": workday_flags,
    }

    # ── Trust score ──
    trust = _compute_trust_score(parsed_rows, tools, operations)

    unique_skus = {r.item_sku for r in parsed_rows}
    unique_machines = {r.resource_code for r in parsed_rows}
    unique_tools = {r.tool_code for r in parsed_rows if r.tool_code}
    workday_count = sum(1 for f in workday_flags if f)

    meta: dict[str, Any] = {
        "rows": len(parsed_rows),
        "machines": len(unique_machines),
        "tools": len(unique_tools),
        "skus": len(unique_skus),
        "dates": n_days,
        "workdays": workday_count,
        "trustScore": trust.score,
        "trustDimensions": trust.dimensions,
        "warnings": warnings,
    }

    missing = [k for k, v in source_columns.items() if not v]
    if missing:
        warnings.append(
            "Colunas não detectadas: "
            + ", ".join(missing)
            + " — serão preenchidas pelo ISOP Mestre ou defaults."
        )

    return ISOPParseResult(
        success=True,
        data=nikufra_data,
        meta=meta,
        source_columns=source_columns,
    )


# ═══════════════════════════════════════════════════════════════
#  Main parser
# ═══════════════════════════════════════════════════════════════


def parse_isop_file(
    filepath_or_bytes: Any,
    *,
    data_only: bool = True,
) -> ISOPParseResult:
    """Parse an ISOP XLSX file and return NikufraData.

    Args:
        filepath_or_bytes: Path string, pathlib.Path, or bytes/BytesIO.
        data_only: If True, read cached values (not formulas).

    Returns:
        ISOPParseResult with success flag, data (NikufraData dict), meta, errors.
    """
    # 1. Load workbook
    try:
        wb = openpyxl.load_workbook(filepath_or_bytes, data_only=data_only)
    except Exception as e:
        logger.exception("Failed to open ISOP XLSX file: %s", e)
        return ISOPParseResult(
            success=False,
            errors=["Ficheiro XLSX inválido — não foi possível abrir."],
        )

    # 2. Find sheet "Planilha1"
    sheet_name: str | None = None
    for sn in wb.sheetnames:
        if sn.lower() == "planilha1":
            sheet_name = sn
            break

    if not sheet_name:
        return ISOPParseResult(
            success=False,
            errors=['Sheet "Planilha1" não encontrada no ficheiro.'],
        )

    ws: Worksheet = wb[sheet_name]

    # 3. Find header row dynamically
    header_result = find_header_row(ws)
    if not header_result:
        return ISOPParseResult(
            success=False,
            errors=[
                'Nenhuma linha de cabeçalho encontrada com "Referência Artigo" e "Máquina" '
                "(procurado nas primeiras 16 linhas)."
            ],
        )

    header_row_idx, headers = header_result  # 1-based row number

    # 4. Map columns by header name
    col_map = build_column_map(headers)
    if not col_map:
        return ISOPParseResult(
            success=False,
            errors=['Colunas "Referência Artigo" e "Máquina" não encontradas nos cabeçalhos.'],
        )

    # 5. Find date columns (after last text column)
    header_cells = list(
        ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=True)
    )[0]

    dates, date_col_indices = find_date_columns(header_cells, col_map)

    if not dates:
        return ISOPParseResult(
            success=False,
            errors=[f"Nenhuma coluna de data encontrada no cabeçalho (linha {header_row_idx})."],
        )

    # 6. Parse workday flags
    workday_flags = find_workday_flags_row(ws, header_row_idx, date_col_indices)
    if workday_flags is None:
        workday_flags = [d.weekday() < 5 for d in dates]
    if len(workday_flags) != len(dates):
        workday_flags = [True] * len(dates)

    # 7. Parse data rows
    data_start_row = header_row_idx + 1  # 1-based
    parsed_rows, warnings = extract_rows(ws, col_map, date_col_indices, data_start_row)

    if not parsed_rows:
        return ISOPParseResult(
            success=False,
            errors=[
                f"Nenhuma linha de dados válida encontrada (a partir da linha {data_start_row})."
            ],
        )

    warnings.append(
        f"Cabeçalho detectado na linha {header_row_idx}, "
        f"dados a partir da linha {data_start_row}, "
        f"{len(dates)} datas, {len(parsed_rows)} operações."
    )

    # 8. Build NikufraData
    source_columns = {
        "hasSetup": col_map.tp_setup >= 0,
        "hasAltMachine": col_map.maq_alt >= 0,
        "hasRate": col_map.pecas_h >= 0,
        "hasParentSku": col_map.produto_acabado >= 0,
        "hasLeadTime": col_map.prz_fabrico >= 0,
        "hasQtdExp": col_map.qtd_exp >= 0,
        "hasTwin": col_map.peca_gemea >= 0,
    }

    return _build_nikufra_data(parsed_rows, dates, workday_flags, warnings, source_columns)
