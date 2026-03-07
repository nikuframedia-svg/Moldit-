# Nikufra Data Service
# Combines ISOP XLSX master data with PP PDF schedule data
# Serves the NikufraPlan frontend component

import hashlib
from pathlib import Path
from typing import Any

from ...core.logging import get_logger
from ..ingest.isop_parser import normalize_code, parse_numeric
from ..ingest.pp_pdf_parser import PPPDFData, parse_pp_pdfs
from .constants import generate_fallback_dates
from .ingest_service import IngestService
from .schemas import NikufraDashboardState

logger = get_logger(__name__)


class NikufraService:
    """Service that parses and combines Nikufra factory data files.

    Reads from:
    - ISOP XLSX: master data (items, machines, tools, routing, rates, stock, setup times)
    - PP PDFs: schedule data (daily quantities, MAN minutes, MO load, ATRASO)

    Output matches the NikufraPlan component's data structures exactly.
    V2: Delegates to IngestService for alerts, stock projections, and trust index.
    """

    def __init__(self, data_dir: Path):
        self.data_dir = data_dir
        self._cache: dict[str, Any] | None = None
        self._cache_hash: str | None = None
        self._live_cache: NikufraDashboardState | None = None
        self._live_cache_hash: str | None = None
        self._known_header_hash: str | None = None

    def _files_hash(self) -> str:
        """Compute SHA-256 hash of source files for cache invalidation."""
        h = hashlib.sha256()
        for name in sorted(self.data_dir.glob("*")):
            h.update(name.name.encode())
            h.update(str(name.stat().st_mtime).encode())
        return h.hexdigest()

    def get_data(self) -> dict[str, Any]:
        """Get combined Nikufra data (V1 format), using cache if available."""
        current_hash = self._files_hash()
        if self._cache and self._cache_hash == current_hash:
            return self._cache

        data = self._build_data()
        self._cache = data
        self._cache_hash = current_hash
        return data

    def get_live_data(self) -> NikufraDashboardState:
        """Get V2 dashboard state with alerts, stock projections, trust index."""
        current_hash = self._files_hash()
        if self._live_cache and self._live_cache_hash == current_hash:
            return self._live_cache

        state = self._build_live_data()
        self._live_cache = state
        self._live_cache_hash = current_hash
        return state

    def reload(self) -> dict[str, Any]:
        """Force re-parse of all source files."""
        self._cache = None
        self._cache_hash = None
        self._live_cache = None
        self._live_cache_hash = None
        return self.get_data()

    def _build_live_data(self) -> NikufraDashboardState:
        """Build V2 dashboard state using IngestService."""
        xlsx_files = list(self.data_dir.glob("ISOP*.xlsx"))
        pg1_files = list(self.data_dir.glob("PP_PG1*.pdf"))
        pg2_files = list(self.data_dir.glob("PP_PG2*.pdf"))

        xlsx_path = xlsx_files[0] if xlsx_files else None

        if not xlsx_path:
            # Return partial state with alert
            ingest = IngestService()
            return ingest.build_dashboard_state(
                isop={"tools": {}, "machines": {}, "items": [], "date_cols": {}},
                pp_data=None,
                xlsx_path=None,
            )

        logger.info(f"Building live data from: {xlsx_path}")
        isop = self._parse_isop(xlsx_path)

        pp_data: PPPDFData | None = None
        try:
            if pg1_files and pg2_files:
                pp_data = parse_pp_pdfs(pg1_files[0], pg2_files[0])
            elif pg1_files:
                from ..ingest.pp_pdf_parser import PPPDFParser

                pp_data = PPPDFParser(pg1_files[0], "PG1").parse()
        except Exception as e:
            logger.error(f"PDF parsing failed: {e}")
            # Will continue with partial state

        ingest = IngestService(known_header_hash=self._known_header_hash)
        state = ingest.build_dashboard_state(
            isop=isop,
            pp_data=pp_data,
            xlsx_path=xlsx_path,
        )

        return state

    def _build_data(self) -> dict[str, Any]:
        """Parse all source files and build the combined data structure."""
        # Find files
        xlsx_files = list(self.data_dir.glob("ISOP*.xlsx"))
        pg1_files = list(self.data_dir.glob("PP_PG1*.pdf"))
        pg2_files = list(self.data_dir.glob("PP_PG2*.pdf"))

        if not xlsx_files:
            raise FileNotFoundError(f"No ISOP XLSX file found in {self.data_dir}")

        xlsx_path = xlsx_files[0]
        logger.info(f"Using ISOP file: {xlsx_path}")

        # Parse ISOP XLSX for master data
        isop = self._parse_isop(xlsx_path)

        # Parse PP PDFs if available
        pp_data: PPPDFData | None = None
        if pg1_files and pg2_files:
            pp_data = parse_pp_pdfs(pg1_files[0], pg2_files[0])
        elif pg1_files:
            from ..ingest.pp_pdf_parser import PPPDFParser

            pp_data = PPPDFParser(pg1_files[0], "PG1").parse()

        # Combine data
        return self._combine(isop, pp_data)

    def _parse_isop(self, xlsx_path: Path) -> dict[str, Any]:
        """Parse ISOP XLSX and return raw master data."""
        import openpyxl

        wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
        if "Planilha1" not in wb.sheetnames:
            raise ValueError("Sheet 'Planilha1' not found in ISOP file")

        ws = wb["Planilha1"]
        header_row = 7  # Row 7 (1-indexed)

        # Build column map from headers
        headers = {}
        for col_idx, cell in enumerate(ws[header_row], start=1):
            if cell.value:
                headers[col_idx] = str(cell.value).strip()

        # Find date columns
        date_cols: dict[int, str] = {}
        for col_idx, header in headers.items():
            if "/" in header and len(header) <= 12:
                # Date column
                date_cols[col_idx] = header

        # Parse data rows
        tools: dict[str, dict[str, Any]] = {}
        machines: dict[str, dict[str, Any]] = {}
        items: list[dict[str, Any]] = []

        for row_idx in range(8, ws.max_row + 1):
            row = [ws.cell(row=row_idx, column=c).value for c in range(1, 20)]

            sku = normalize_code(row[3]) if row[3] else None  # Referência Artigo
            if not sku:
                continue

            customer = normalize_code(row[0])
            customer_name = str(row[1]).strip() if row[1] else ""
            parent_sku = normalize_code(row[2])
            name = str(row[4]).strip() if row[4] else ""
            lot_qty = int(parse_numeric(row[5], 0) or 0)
            machine = normalize_code(row[6])
            alt_machine = normalize_code(row[7])
            tool = normalize_code(row[8])
            setup_time = parse_numeric(row[9], 0) or 0
            rate = int(parse_numeric(row[10], 0) or 0)
            operators = int(parse_numeric(row[11], 1) or 1)
            # Col 12 = Qtd Exp (stock/expedição)

            # Get stock value from Qtd Exp column
            stock = int(parse_numeric(row[12], 0) or 0) if len(row) > 12 else 0

            if machine and machine not in machines:
                # Determine area from PP data or heuristic
                area = "PG1"  # Default, will be overridden by PP data
                machines[machine] = {
                    "id": machine,
                    "area": area,
                    "man": [0] * 8,
                }

            if tool and tool not in tools:
                tools[tool] = {
                    "id": tool,
                    "m": machine or "",
                    "alt": alt_machine or "-",
                    "s": setup_time / 60.0 if setup_time else 0,  # Convert min to hours
                    "pH": rate,
                    "op": operators,
                    "skus": [],
                    "nm": [],
                    "lt": lot_qty,
                    "stk": 0,
                }

            # Add SKU to tool
            if tool and tool in tools:
                if sku not in tools[tool]["skus"]:
                    tools[tool]["skus"].append(sku)
                    tools[tool]["nm"].append(name)
                # Update stock
                if stock > 0:
                    tools[tool]["stk"] = stock

            items.append(
                {
                    "sku": sku,
                    "name": name,
                    "customer": customer,
                    "customer_name": customer_name,
                    "tool": tool,
                    "machine": machine,
                    "alt_machine": alt_machine,
                    "lot_qty": lot_qty,
                    "rate": rate,
                    "setup_time": setup_time,
                    "operators": operators,
                    "stock": stock,
                }
            )

        wb.close()

        return {
            "tools": tools,
            "machines": machines,
            "items": items,
            "date_cols": date_cols,
        }

    def _combine(
        self,
        isop: dict[str, Any],
        pp_data: PPPDFData | None,
    ) -> dict[str, Any]:
        """Combine ISOP master data with PP schedule data."""
        tools_list = list(isop["tools"].values())

        # Use PP data for dates, machines, operations, MO
        dates, days_label = generate_fallback_dates()
        mo: dict[str, list[float]] = {"PG1": [0] * 8, "PG2": [0] * 8}

        machines_list = []
        ops_list = []
        machine_area_map: dict[str, str] = {}

        if pp_data:
            dates = pp_data.dates or dates
            days_label = pp_data.days_label or days_label
            mo = pp_data.mo_load if pp_data.mo_load else mo

            # Build machines from PP data
            for mb in pp_data.machines:
                machines_list.append(
                    {
                        "id": mb.machine_id,
                        "area": mb.area,
                        "man": mb.man_minutes,
                    }
                )
                machine_area_map[mb.machine_id] = mb.area

                # Build operations
                for i, op in enumerate(mb.operations):
                    # Cross-reference with ISOP tool data for setup time
                    isop_tool = isop["tools"].get(op.tool_code, {})
                    setup_h = isop_tool.get("s", 0) if isop_tool else 0
                    if op.setup_hours > 0:
                        setup_h = op.setup_hours

                    ops_list.append(
                        {
                            "id": f"OP{len(ops_list) + 1:02d}",
                            "m": op.machine,
                            "t": op.tool_code,
                            "sku": op.sku,
                            "nm": op.name,
                            "pH": op.pcs_per_hour or isop_tool.get("pH", 0),
                            "atr": op.atraso,
                            "d": op.daily_qty,
                            "s": setup_h,
                            "op": op.operators or isop_tool.get("op", 1),
                        }
                    )

            # Update tool stock from PP data
            for op in pp_data.all_operations:
                if op.tool_code in isop["tools"] and op.stock > 0:
                    isop["tools"][op.tool_code]["stk"] = op.stock

        # If no PP data, create machines from ISOP
        if not machines_list:
            for mid, mdata in isop["machines"].items():
                machines_list.append(mdata)

        # Update machine areas from PP data
        for m in machines_list:
            if m["id"] in machine_area_map:
                m["area"] = machine_area_map[m["id"]]

        # Update ISOP tools with PP-derived machine areas
        for tool in tools_list:
            if tool["m"] in machine_area_map:
                pass  # Area info stays with machine

        # MOCK: Event history — hardcoded for frontend development.
        # Production: query from domain.run_events.service
        history = [
            {
                "dt": "01/02",
                "type": "machine_down",
                "mach": "PRM039",
                "tool": "BFP092",
                "action": "BFP092 → PRM043",
                "result": "Retomada 45min",
                "roi": "—",
            },
            {
                "dt": "30/01",
                "type": "maintenance",
                "mach": "PRM031",
                "tool": "BFP079",
                "action": "Manutenção preventiva",
                "result": "Sem impacto",
                "roi": "—",
            },
            {
                "dt": "28/01",
                "type": "urgent_order",
                "mach": "PRM019",
                "tool": "BFP080",
                "action": "Resequenciamento",
                "result": "OTD 100%",
                "roi": "—",
            },
            {
                "dt": "27/01",
                "type": "operator",
                "mach": "PRM043",
                "tool": "BFP172",
                "action": "Pool Y reassignado",
                "result": "Delay 30min T1",
                "roi": "—",
            },
            {
                "dt": "25/01",
                "type": "machine_down",
                "mach": "PRM031",
                "tool": "BFP114",
                "action": "BFP114 → PRM039",
                "result": "Setup +1.25h ok",
                "roi": "—",
            },
            {
                "dt": "23/01",
                "type": "machine_down",
                "mach": "PRM039",
                "tool": "BFP178",
                "action": "BFP178 → PRM043",
                "result": "Sem impacto",
                "roi": "—",
            },
            {
                "dt": "20/01",
                "type": "maintenance",
                "mach": "PRM043",
                "tool": "BFP202",
                "action": "Corretiva 2h",
                "result": "Sem alt. dispo.",
                "roi": "—",
            },
        ]

        return {
            "dates": dates,
            "days_label": days_label,
            "mo": mo,
            "machines": machines_list,
            "tools": tools_list,
            "operations": ops_list,
            "history": history,
        }
